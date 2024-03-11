import os
import shutil
from openpyxl import load_workbook
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from globalv import *

class AnalizadorFechas():
    # Definir la constante estilo_celda a nivel de clase
    estilo_celda = "border: 1px solid #dddddd; text-align: left; padding: 8px;"

    def __init__(self, citas_folder="citas", update_folder="citas/update"):
        self.citas_folder = citas_folder
        self.update_folder = update_folder
        self.email_sender = 'hervosoisaac@gmail.com'
        self.email_password = 'wptl madg hrjo yfwt'
        self.email_receivers = email_receivers

    def leer_fechas_excel(self, file_path):
        try:
            fechas = []
            workbook = load_workbook(filename=file_path)
            sheet = workbook.active
            for row in sheet.iter_rows(values_only=True):
                fechas.extend(row)
            return fechas
        except Exception as e:
            print(f"Error al leer fechas del archivo {file_path}: {e}")
            return []

    def enviar_correo_error(self, error_message):
        try:
            # Configurar el servidor SMTP de Gmail
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls()
            server.login(self.email_sender, self.email_password)

            # Crear mensaje
            msg = MIMEMultipart()
            msg['From'] = self.email_sender
            msg['To'] = ', '.join(self.email_receivers)
            msg['Subject'] = 'Error en el bot de Cosevi'

            # Cuerpo del mensaje
            body = f"Se ha producido un error en el bot de Cosevi:\n\n{error_message}"
            msg.attach(MIMEText(body, 'plain'))

            # Enviar correo electrónico
            server.sendmail(self.email_sender, self.email_receivers, msg.as_string())
            print("Correo electrónico de error enviado correctamente.")

            # Cerrar conexión
            server.quit()
        except Exception as e:
            print(f"Error al enviar correo electrónico de error: {e}")
    def enviar_correo(self, subject, body, sede, cita_proxima):
        try:
            # Configurar el servidor SMTP de Gmail
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls()
            server.login(self.email_sender, self.email_password)

            # Crear mensaje en formato HTML
            msg = MIMEMultipart()
            msg['From'] = self.email_sender
            msg['Subject'] = subject

            # Estilos CSS
            estilo_letras = "font-family: Arial, sans-serif; font-size: 16px; color: #333;"
            estilo_tabla = "border-collapse: collapse; width: 100%;"
            button_style = "background-color: #4CAF50; color: white; padding: 12px 20px; border: none; border-radius: 4px; cursor: pointer; text-decoration: none;"

            # Adjuntar cuerpo del mensaje como HTML
            mensaje_html = f"""
                    <html>
                        <head></head>
                        <body style="{estilo_letras}">
                            <h2 style="{estilo_letras}">Se encontraron citas extraordinarias en la sede: {sede}</h2>
                            <table style="{estilo_tabla}">
                                <tr>
                                    <th style="{self.estilo_celda}">Fechas</th>
                                </tr>
                                {body}
                            </table>
                            <hr>
                            <p style="{estilo_letras}">La cita más próxima es:</p>
                            <h3 style="{estilo_letras}">{cita_proxima}</h3>
                            <form action="https://servicios.educacionvial.go.cr/Formularios/IngresarCuenta">
                                <input type="submit" value="Ir a COSEVI" style="{button_style}">
                            </form>
                        </body>
                    </html>
                    """
            msg.attach(MIMEText(mensaje_html, 'html'))

            # Agregar destinatarios al campo "To"
            msg['To'] = ', '.join(self.email_receivers)

            # Enviar correo electrónico
            server.sendmail(self.email_sender, self.email_receivers, msg.as_string())

            print("Correo electrónico enviado exitosamente.")

            # Cerrar conexión
            server.quit()
        except Exception as e:
            print(f"Error al enviar correo electrónico: {e}")

    def actualizar_carpeta_update(self):
        try:
            # Copiar archivos de citas a update
            for filename in os.listdir(self.citas_folder):
                if filename.endswith(".xlsx"):
                    citas_path = os.path.join(self.citas_folder, filename)
                    update_path = os.path.join(self.update_folder, filename)

                    shutil.copy(citas_path, update_path)  # Copiar de citas a update
            print("La carpeta 'update' ha sido actualizada exitosamente.")
        except Exception as e:
            print(f"Error al actualizar la carpeta 'update': {e}")

    def reemplazar_archivos(self):
        try:
            for filename in os.listdir(self.update_folder):
                if filename.endswith(".xlsx"):
                    update_path = os.path.join(self.update_folder, filename)
                    citas_path = os.path.join(self.citas_folder, filename)
                    fechas_update = self.leer_fechas_excel(update_path)

                    if os.path.exists(citas_path):
                        fechas_citas = self.leer_fechas_excel(citas_path)

                        if fechas_update != fechas_citas:
                            print(f"El archivo {filename} en 'update' es diferente al de 'citas'.")
                            sede = filename.split("_")[0].upper()

                            if len(fechas_citas) > 3:  # Verificar si hay suficientes fechas en la lista
                                cita_proxima = fechas_citas[3]
                            else:
                                cita_proxima = "No hay citas disponibles"

                            fechas_nuevas = [fecha for fecha in fechas_citas if fecha not in fechas_update]
                            if fechas_nuevas:
                                body = ""
                                for fecha in fechas_nuevas:
                                    body += f"<tr><td style='{self.estilo_celda}'>{fecha}</td></tr>"
                                # Buscar la cita más próxima
                                # Enviar correo con citas extraordinarias y la cita más próxima
                                self.enviar_correo(f"({tipo_cita}) {sede} Nuevas citas extraordinarias", body, sede,
                                                   cita_proxima)
                            else:
                                print("No se encontraron citas extraordinarias nuevas.")
                    else:
                        print(f"No se encontró el archivo {filename} en la carpeta 'citas'.")
            print(f"Se analizaron todas las cedes")

            # Copiar archivos de citas a update
            self.actualizar_carpeta_update()

        except Exception as e:
            print(f"Error al reemplazar archivos: {e}")


analizador = AnalizadorFechas()
analizador.reemplazar_archivos()
