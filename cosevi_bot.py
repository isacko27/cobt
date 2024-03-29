from openpyxl.reader.excel import load_workbook
from undetected_chromedriver import Chrome, ChromeOptions
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException
import pyautogui
import time
import re
import os
from openpyxl import Workbook
import random
from analizar_fechas import analizador
from globalv import tipo_cita

class CoseviBot:
    def __init__(self, driver=None):
        self.anos_fechas = {}
        self.cede = ""
        self.driver = driver

    def set_driver(self, driver):
        self.driver = driver
    def simular_scroll_y_mouse(self):
        try:
            # Simular scroll
            scroll_amount = random.randint(5, 20)  # Genera un valor aleatorio entre 5 y 20
            pyautogui.scroll(scroll_amount)

            # Simular movimiento del mouse
            screen_width, screen_height = pyautogui.size()
            center_x = screen_width // 2
            center_y = screen_height // 2
            move_x = random.randint(-100, 100)  # Genera un valor aleatorio entre -100 y 100
            move_y = random.randint(-100, 100)  # Genera un valor aleatorio entre -100 y 100
            pyautogui.moveTo(center_x + move_x, center_y + move_y, duration=random.uniform(0.1, 0.5))  # Duración aleatoria entre 0.1 y 0.5 segundos

        except Exception as e:
            print("Error al simular scroll, movimiento del mouse o clic:", e)

    def debug(self):
        try:
            print("Iniciando modo de depuración...")

            # Esperar 2 minutos
            time.sleep(1000)

            print("Continuando con el código...")
            # Continuar con el resto del código aquí...

        except Exception as e:
            print("Error en el modo de depuración:", e)

    def cerrar_sesion(self):
        try:
            # Buscar y hacer clic en el enlace "Salir"
            enlace_salir = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located(
                    (By.XPATH, "//a[@class='encabezado-nombre-usuario encabezado-boton-salir']"))
            )
            enlace_salir.click()
            print("Se cerró la sesión correctamente.")
        except Exception as e:
            print("Error al cerrar sesión:", e)

    def esperar_modal_desaparezca(self):
        try:
            WebDriverWait(self.driver, 10).until(
                EC.invisibility_of_element_located((By.CSS_SELECTOR, ".modal-carga"))
            )
            time.sleep(2)
        except Exception as e:
            print("Error al esperar que el modal de carga desaparezca:", e)

    def IniciarSesion(self, identificacion, contrasena, tipo_identificacion, max_intentos=3):
        intentos = 0
        while intentos < max_intentos:
            try:
                self.driver.get('https://servicios.educacionvial.go.cr/Formularios/IngresarCuenta')
                selector_tipo_identificacion = self.driver.find_element(By.CLASS_NAME, 'selector-tipos-identificacion')
                Select(selector_tipo_identificacion).select_by_value(tipo_identificacion)
                WebDriverWait(self.driver, 60).until(EC.presence_of_element_located((By.ID, "identificacion")))
                identificacion_input = self.driver.find_element(By.ID, 'identificacion')
                identificacion_input.send_keys(identificacion)
                WebDriverWait(self.driver, 60).until(EC.presence_of_element_located((By.ID, "contrasena")))
                contrasena_input = self.driver.find_element(By.ID, 'contrasena')
                contrasena_input.send_keys(contrasena)
                boton_acceder = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.ID, 'botonAcceder')))
                boton_acceder.click()
                self.esperar_modal_desaparezca()

                # Buscar y hacer clic en el botón "Si" si está presente
                try:
                    boton_si = self.driver.find_element(By.CLASS_NAME, "cancel")
                    boton_si.click()
                    print("Se encontró y clickeó el botón 'Si'")
                except:
                    pass  # Si no se encuentra el botón "Si", simplemente continúa con el programa

                # Esperar hasta que aparezca el enlace "Salir"
                WebDriverWait(self.driver, 10).until(EC.presence_of_element_located(
                    (By.XPATH, "//a[@class='encabezado-nombre-usuario encabezado-boton-salir']")))
                print("Se inició sesión correctamente")
                return

            except Exception as e:
                print("Error al iniciar sesión:", e)
                intentos += 1
                if intentos < max_intentos:
                    print("Volviendo a intentar...")
                    time.sleep(5)
                    continue
                else:
                    print("Se alcanzó el número máximo de intentos. No se pudo iniciar sesión.")
                    break
        # Si se superó el número máximo de intentos, se vuelve a intentar la función de inicio de sesión
        self.IniciarSesion(identificacion, contrasena, max_intentos)

    def ingresarRecibo(self, numero_recibo, clase_licencia):
        max_intentos = 3
        intento = 0

        while intento < max_intentos:
            intento += 1
            try:
                self.driver.get('https://servicios.educacionvial.go.cr/Formularios/MatriculaPruebaPractica')

                numero_recibo_input = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.ID, 'numRecibo')))
                numero_recibo_input.send_keys(numero_recibo)
                print("Se colocó el número del recibo")
                # self.simular_scroll_y_mouse()

                select_element = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, 'select.form-control.selector-licencia'))
                )
                select = Select(select_element)
                select.select_by_value(clase_licencia)

                checkbox = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.ID, 'check-AceptaTerminos')))
                checkbox.click()
                print("Se aceptaron los términos")

                # self.simular_scroll_y_mouse()
                submit_button = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.ID, 'botonContinuar')))
                # self.simular_scroll_y_mouse()
                submit_button.click()
                print("Se consultaron las Cedes")
                self.esperar_modal_desaparezca()
                return

            except Exception as e:
                print("Error al ingresar el recibo:", e)
                time.sleep(1)
                continue

        print(f"Se superó el número máximo de intentos ({max_intentos}).")
        return "Se superó el número máximo de intentos"

    def consultarCede(self, nombre_cede):
        self.cede = nombre_cede
        try:
            WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.ID, "listaSedes")))
            cedes = self.driver.find_elements(By.XPATH, "//ul[@id='listaSedes']/li")
            cede_encontrada = None
            print(f"Entrando en Cede: {nombre_cede}")
            self.esperar_modal_desaparezca()

            for cede in cedes:
                if nombre_cede.lower() in cede.text.lower():
                    cede_encontrada = cede
                    break

            if cede_encontrada:
                # self.simular_scroll_y_mouse()
                cede_encontrada.click()
                print(f"Se seleccionó la cede: {nombre_cede}")

                boton_continuar = self.driver.find_element(By.ID, "botonContinuarSedes")
                boton_continuar.click()
                print("Se hizo clic en 'Continuar'")
                time.sleep(2)
                self.esperar_modal_desaparezca()

                citas_disponibles = self.verificarCitasDisponibles()
                if citas_disponibles:
                    print(f"[+]Hay citas disponibles en {nombre_cede}")
                    self.obtenerFechasDisponibles()
                    self.guardarFechasEnExcel()
                else:
                    print(f"[-]NO hay citas en {nombre_cede}")
                    self.verificar_mensaje_sin_citas()

            else:
                print(f"No se encontró la cede: {nombre_cede}")

        except Exception as e:
            print("Error al consultar la cede:", e)
        analizador.reemplazar_archivos()

    def verificarCitasDisponibles(self):
        try:
            lista_errores = WebDriverWait(self.driver, 5).until(
                EC.presence_of_element_located((By.ID, "listaErrores"))
            )
            if lista_errores.find_elements(By.XPATH, "./*"):
                return False
            else:
                return True

        except NoSuchElementException:
            return True

        except Exception as e:
            print("Error al verificar citas disponibles:", e)
            return False

    def verificar_mensaje_sin_citas(self):
        try:
            mensaje_sin_citas = WebDriverWait(self.driver, 5).until(
                EC.presence_of_element_located((By.ID, "CD"))
            )
            if mensaje_sin_citas.text.strip() == "En ésta regional no existen espacios disponibles, favor intentar más tarde o seleccionar otra":
                print("Se encontró el mensaje de 'Sin citas disponibles'. Vaciar el archivo Excel...")
                self.vaciar_excel()
            else:
                print("No se encontró el mensaje de 'Sin citas disponibles'.")
        except NoSuchElementException:
            print("No se encontró el mensaje de 'Sin citas disponibles'.")
        except Exception as e:
            print("Error al verificar mensaje de 'Sin citas disponibles':", e)

    def vaciar_excel(self):
        try:
            file_path = os.path.join("citas", f"{self.cede}_citas_disponibles.xlsx")
            if os.path.exists(file_path):
                wb = load_workbook(file_path)
                ws = wb.active
                ws.delete_rows(2, ws.max_row)  # Eliminar todas las filas excepto la primera
                wb.save(file_path)
                print(f"Se vació el archivo Excel '{file_path}'.")
            else:
                print(f"No se encontró el archivo Excel '{file_path}'.")
        except Exception as e:
            print("Error al vaciar el archivo Excel:", e)
    def obtenerFechasDisponibles(self):
        try:
            citas = WebDriverWait(self.driver, 10).until(
                EC.presence_of_all_elements_located((By.XPATH, "//div[@class='jqx-grid-cell-left-align']"))
            )

            fechas = []
            for cita in citas:
                fecha = cita.text
                fechas.append(fecha)

            if fechas:
                print("Citas disponibles los días:")
                fechas_limpias = self.limpiarFechas(fechas)
                fechas_anio = self.agruparPorAnio(fechas_limpias)
                self.mostrarFechas(fechas_anio)
                try:
                    boton_atras = WebDriverWait(self.driver, 10).until(
                        EC.presence_of_element_located((By.ID, 'botonAtrasCitas'))
                    )
                    boton_atras.click()
                    print("Volviendo atrás...")
                except Exception as e:
                    print("Error al hacer clic en el botón Atrás:", e)
            else:
                print("No se encontraron fechas disponibles")

        except Exception as e:
            print("Error al obtener fechas disponibles:", e)

    def limpiarFechas(self, fechas):
        patron_fecha = r'\b(?:Lunes|Martes|Miércoles|Jueves|Viernes|Sábado|Domingo)\s\d{2}/\d{2}/\d{4}\b'
        fechas_limpias = []

        for fecha in fechas:
            coincidencias = re.findall(patron_fecha, fecha)
            fechas_limpias.extend(coincidencias)

        return fechas_limpias

    def agruparPorAnio(self, fechas):
        fechas_anio = {}
        for fecha in fechas:
            anio = fecha.split('/')[-1]
            if anio not in fechas_anio:
                fechas_anio[anio] = []
            fechas_anio[anio].append(fecha)

        self.anos_fechas.update(fechas_anio)
        return fechas_anio

    def mostrarFechas(self, fechas_anio):
        for anio, fechas in fechas_anio.items():
            print(f"Año {anio}:")
            for fecha in fechas:
                print(fecha)
        self.guardarFechasEnExcel()

    def guardarFechasEnExcel(self):
        try:
            wb = Workbook()
            ws = wb.active
            ws.append(["Año", "Fecha"])

            for anio, fechas in self.anos_fechas.items():
                for fecha in fechas:
                    ws.append([anio, fecha])

            file_path = os.path.join("citas", f"{self.cede}_citas_disponibles.xlsx")
            wb.save(file_path)
            print(f"Se guardaron las fechas de citas disponibles en '{file_path}'.")

        except Exception as e:
            print("Error al guardar las fechas en Excel:", e)

    def CerrarSesion(self):
        try:
            print("Saliendo del programa...")
            if self.driver:
                self.driver.quit()
                print("Se cerró el driver correctamente.")
            else:
                print("No hay instancia del driver para cerrar.")
        except Exception as e:
            print("Error al cerrar el driver:", e)

    def Restart(self):
        try:
            print("Reiniciando...")
            if self.driver:
                self.driver.get("https://google.com")
            else:
                print("No hay instancia del driver para reiniciar.")
        except Exception as e:
            print("Error al reiniciar el driver:", e)
